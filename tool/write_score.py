from openpyxl import load_workbook
import csv
import os

def update_excel_and_save_csv(excelA_path, excelB_path, output_csv_path,update_dir_all_path):
    update_dir_all_path = os.path.join(update_dir_all_path, output_csv_path)
    # 載入 ExcelA 和 ExcelB 的工作簿與工作表
    wbA = load_workbook(excelA_path)
    wsA = wbA.active

    wbB = load_workbook(excelB_path)
    wsB = wbB.active

    # 建立 ExcelA 的 A:D 資料對應（從第2列開始）
    mapping = {}
    for row in wsA.iter_rows(min_row=2, max_col=4, values_only=False):
        key = row[0].value  # A欄
        value = row[3].value  # D欄
        mapping[key] = value
        print(f"Mapping: {key} -> {value}")  # Debugging line
    #print(f"Mapping dict: {mapping}")  # Debugging line

    # 更新 ExcelB 的 G欄資料（第7欄）
    for row in wsB.iter_rows(min_row=2, max_col=1):
        cell = row[0]  # A欄
        if cell.value in mapping:
            wsB.cell(row=cell.row, column=7, value=mapping[cell.value])  # 寫入 G欄

    # 檢查 G欄是否為 "-"，若是則改為 0
    for row in wsB.iter_rows(min_row=2, min_col=7, max_col=7):  # G欄是第7欄
        cell = row[0]
        if cell.value == "-":
            cell.value = 0

    # 將更新後的 ExcelB 寫出為 CSV
    with open(update_dir_all_path, mode='w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        for row in wsB.iter_rows(values_only=True):
            writer.writerow(row)

    print(f"已更新並轉換為 CSV: {output_csv_path}")


if __name__ == '__main__':
    base_dir = os.getcwd()
    excelA_path = 'Score_03172.xlsx'
    excelB_path = 'testing.xlsx'
    output_csv_path = 'updated_excelB.csv'
    update_excel_and_save_csv(excelA_path, excelB_path, output_csv_path,base_dir)