from a_txt2pdf import *
import os
import re
import subprocess
import time
import zipfile
import shutil
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

compile_path = r"C:\msys64\mingw64\bin\g++.exe"
test_input_file = "testinput.txt"
total_file_name = "total.txt"
error_student_folder = "0Error"
file_extension='.cpp'
max_retries = 3     # 最多重試次數
retry_delay = 1     # 每次重試間隔秒數
test_timeout= 10    # 每個程式最大執行時間(s)
test_file_dir = "test_file"
hw_dir= "HW_folder"

'''
test input file name: 1input.txt, 2input.txt, ...
answer file name: 1ans.txt, 2ans.txt, ...
testinput.txt: 暫存每次測試的輸入
total.txt: 統計學生的答題狀況
Score_{作業日期}.xlsx: 紀錄該次作業狀況
file_extension: 檔案副檔名
test_file_dir: 測試檔案資料夾
hw_dir: 作業資料夾

'''


#解壓縮(.zip)
def unzip_file(base_dir):
    hw_folder = os.path.join(base_dir, hw_dir)
    files = os.listdir(hw_folder)

    zip_files = [f for f in files if f.lower().endswith(".zip")]

    if len(zip_files) == 1:
        zip_path = os.path.join(hw_folder, zip_files[0])
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(hw_folder)
        os.remove(zip_path)
        print(f"已解壓縮 {zip_files[0]} 並刪除原壓縮檔")
    elif len(zip_files) == 0:
        print(f"錯誤：{hw_folder} 內沒有找到壓縮檔")
    else:
        print(f"錯誤：{hw_folder} 內有多個壓縮檔，請手動確認")
#generate excel
def generate_excel(check_excel,score):
    """
    生成一個 Excel 檔案 (StudentList.xlsx) 存放學生資料夾名稱與學號。
    Excel 的 A1 儲存格填入 "dir_name"，B1 填入 "S_id"，
    從 A2 開始，每一列依序填入學生資料夾的新名稱與對應的學號（以底線分隔，第一部分為學號）。
    """
    if check_excel and os.path.exists(excel_file):
        # 開啟既有 Excel
        wb = load_workbook(excel_file)
        ws = wb.active  # 假設處理第一個工作表

        # 收集現有的 S_id（B 欄）到一個 set
        existing_sids = set()
        for row_idx in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=2).value
            if cell_value:
                existing_sids.add(cell_value)
        
        # 從下一個空白列開始插入新資料
        row = ws.max_row + 1
        print(f"已讀取 {excel_file}，目前已有 {len(existing_sids)} 筆學號紀錄。")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Student List"
        ws["A1"] = "Student ID"
        ws["B1"] = "Student name"
        ws["C1"] = "Error count"
        ws["D1"] = "Score"
        ws["E1"] = "Note"
        ws["F1"] = f"Points: {score}"

        existing_sids = set()
        row = 2
        print(f"建立新的 {excel_file} 檔案")
    wb.save(excel_file)
    print(f"Excel 檔案已儲存：{excel_file}")
#rename student folders and update excel
def student_folder_name_excel(hw_dir_path,check_excel):
    wb = load_workbook(excel_file)
    ws = wb.active  # 假設處理第一個工作表
    row = 2
    for folder in os.listdir(hw_dir_path):
        folder_path = os.path.join(hw_dir_path, folder)
        chinese_name = None
        student_id = None
        if os.path.isdir(folder_path):
            # 修改正則表達式以捕獲中文姓名與學號
            #m = re.search(r'^([\u4e00-\u9fff]+).*?(\d{8,9}[A-Za-z])(?=_|$)', folder)
            m = re.search(r'^([\u4e00-\u9fff]+).*?(\d{8}[A-Za-z])', folder)

            #print("m: ",m)
            if m:
                chinese_name = m.group(1)
                student_id = m.group(2)
                print(f"chinse_name: {chinese_name}, student_id: {student_id}")
                #new_name = f"{chinese_name}_{student_id}"
                #new_name = f"{student_id}_{chinese_name}"
                new_name = f"{student_id}"
                new_path = os.path.join(hw_dir_path, new_name)
                if folder != new_name:
                    if os.path.exists(new_path):
                        print(f"新名稱 '{new_name}' 已存在，無法重新命名 '{folder}'")
                    else:
                        print(f"將資料夾 '{folder}' 重新命名為 '{new_name}'")
                        os.rename(folder_path, new_path)
                else:
                    print(f"資料夾 '{folder}' 已符合格式，不做修改。")
            else:
                print(f"資料夾 '{folder}' 中找不到符合格式的中文姓名及學號，跳過重新命名。")
            
            
            if not check_excel:
                # 收集現有的 S_id（B 欄）到一個 set
                existing_sids = set()
                # 假設新名稱格式為 "學號_中文姓名"，底線前為學號
                '''
                dir_name = folder
                parts = folder.split("_")
                s_id = parts[0] if parts else ""
                '''
                s_id=chinese_name

                if s_id==error_student_folder:
                    continue
                # 若該學號不在 Excel 紀錄中，則插入新資料
                elif s_id and s_id not in existing_sids:
                    ws.cell(row=row, column=1, value=student_id)  # A 欄：dir_name
                    ws.cell(row=row, column=2, value=s_id)      # B 欄：S_id
                    row += 1
                    existing_sids.add(s_id)
    if not check_excel:
        # 1. 讀取從第 2 列開始的所有資料（含多欄）
        rows_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # row 會是一個 tuple，包含該列所有欄位的值
            rows_data.append(row)

        # 2. 依照 A 欄（即 row[0]）進行排序
        #    假設 A 欄存放的是字串，如 "11303114A" 等
        rows_data.sort(key=lambda x: x[0])

        # 3. 清除原先從第 2 列開始的內容
        max_row = ws.max_row
        max_col = ws.max_column
        for r in range(2, max_row + 1):
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c, value=None)

        # 4. 寫回排序後的資料到第 2 列起
        for i, row_tuple in enumerate(rows_data, start=2):
            for j, cell_value in enumerate(row_tuple, start=1):
                ws.cell(row=i, column=j, value=cell_value)
    wb.save(excel_file)
    print(f"Excel 檔案已儲存：{excel_file}")


#move non cpp files
def move_non_cpp_folders(hw_folder_path):
    check_count = 1
    error_folder_path = os.path.join(hw_folder_path, error_student_folder)

    # 如果錯誤資料夾不存在，則建立它
    if not os.path.exists(error_folder_path):
        os.makedirs(error_folder_path)

    # 遍歷 HW_folder 內的學生資料夾
    for student_folder in os.listdir(hw_folder_path):
        student_folder_path = os.path.join(hw_folder_path, student_folder)

        # 跳過錯誤資料夾本身
        if student_folder == error_student_folder:
            continue

        # 確保是資料夾
        if os.path.isdir(student_folder_path):
            has_non_cpp = False  # 用來判斷該學生資料夾內是否有非 .cpp 檔案

            for file in os.listdir(student_folder_path):
                file_path = os.path.join(student_folder_path, file)

                # 檢查是否為檔案，且副檔名不是 .cpp
                if os.path.isfile(file_path) and not file.lower().endswith(file_extension):
                    has_non_cpp = True
                    break  # 找到非 .cpp 檔案就可提前結束該資料夾的搜尋

            # 如果該學生資料夾內含有非 .cpp 檔案，則移動至錯誤資料夾
            if has_non_cpp:
                target_path = os.path.join(error_folder_path, student_folder)
                
                # 如果目標資料夾已存在（避免衝突），先刪除它
                if os.path.exists(target_path):
                    shutil.rmtree(target_path)

                shutil.move(student_folder_path, target_path)
                print(f"{check_count}. 已將 {student_folder} 移動到 {error_student_folder} 資料夾")
                check_count += 1


#檢查測試檔案是否存在
def check_test_files(num_programs):
    if not os.path.exists(test_file_dir):
        print("未找到測試程式，正在建立 test_file 資料夾...")
        os.makedirs(test_file_dir)  # 建立資料夾
        print("test_file 資料夾已建立。")
        return True
    else:
        print("test_file 資料夾已存在，繼續執行程式。")

    for i in range(1, num_programs + 1):
        input_filename = f"{i}input.txt"
        ans_filename = f"{i}ans.txt"
        input_file_path = os.path.join(test_file_dir, input_filename)
        ans_file_path = os.path.join(test_file_dir, ans_filename)
        if not os.path.isfile(input_file_path):
            print(f"找不到測試檔 {input_filename}，請檢查。")
            return True
        if not os.path.isfile(ans_file_path):
            print(f"找不到答案檔 {ans_filename}，請檢查。")
            return True
    return False

def read_blocks(file_path):
    blocks = []
    current_block = []
    with open(file_path, "r") as f:
        for line in f:
            line = line.rstrip("\n")
            if line.strip() == "":
                if current_block:
                    blocks.append(current_block)
                    current_block = []
            else:
                current_block.append(line)
        if current_block:
            blocks.append(current_block)
    return blocks
#compile and test
def process_student_folder(folder, num_programs,score,base_dir):
    st_info = []
    msg = ""
    total_score = 0
    results = []
    error_count = 0
    wrong_questions = []  # 記錄哪幾題錯誤
    for i in range(1, num_programs + 1):
        cpp_filename = f"{i}.cpp"
        cpp_path = os.path.join(folder, cpp_filename)
        test_file_path = os.path.join(base_dir, test_file_dir)

        msg="\n------------------------------------------"
        print(f"{msg}")
        st_info.append(f"{msg}")

        if not os.path.isfile(cpp_path):
            msg = f"*找不到 {cpp_filename},Pass {i}"
            print(f"{msg}")
            st_info.append(f"{msg}")
            results.append("X")
            error_count += 1
            wrong_questions.append(i)
            continue

        # 編譯程式，將執行檔命名為「i」(不含副檔名)
        exe_path = os.path.join(folder, f"{i}")
        #compile_cmd = ["g++", cpp_path, "-o", exe_path]
        compile_cmd = [compile_path,"-o", exe_path,cpp_path]
        #compile_cmd = [compile_path, cpp_path, "-o", exe_path]


        result = subprocess.run(compile_cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        if result.returncode != 0:
            msg=f"編譯 {os.path.basename(cpp_path)} 時發生錯誤: {result.stderr}"
            print(f"{msg}")
            st_info.append(f"{msg}")
            results.append("X")
            error_count += 1
            wrong_questions.append(i)
            continue
        else:
            msg=f"成功編譯 {os.path.basename(cpp_path)} 為 {os.path.basename(exe_path)}.exe"
            print(f"{msg}")
            st_info.append(f"{msg}")

        # 設定測試檔案：
        input_filename = f"{i}input.txt"
        input_file_path = os.path.join(test_file_path, input_filename)
        
        ans_filename = f"{i}ans.txt"
        ans_file_path = os.path.join(test_file_path, ans_filename)

        # 輸出結果
        output_filename = f"{i}output.txt"
        output_file_path = os.path.join(folder, output_filename)

        # clear output_file_path
        with open(output_file_path, "w") as outf:
            outf.write("")

        # read input_file_path
        if not os.path.isfile(input_file_path):
            msg=f"找不到輸入檔 {input_file_path}，跳過 {cpp_filename}。"
            print(f"{msg}")
            st_info.append(f"{msg}")
            results.append("X")
            error_count += 1
            wrong_questions.append(i)
            continue

        input_blocks = read_blocks(input_file_path)
        ans_blocks = read_blocks(ans_file_path)
        '''
        with open(input_file_path, "r") as f:
            #lines = f.readlines()
            lines = [line.rstrip() for line in f if line.strip() != ""]
        f.close()
        for line in lines:
            test_data = line.strip()
            if test_data == "":
                continue  # 忽略空行
        '''

        ans_check = 0   #紀錄答對數
        total_count = len(input_blocks) #總筆數
        ans_count = len(ans_blocks)  #答案數
        if total_count != ans_count:
            msg=f"測試案例數量與答案數量不符，測試案例數量: {total_count}，答案數量: {ans_count}"
            print(f"{msg}")
            st_info.append(f"{msg}")
            er=input("是否繼續? (y/n): ")
            if er.lower() == "n":
                return
            continue

        for idx, block in enumerate(input_blocks):
            test_data = "\n".join(block) + "\n"
            
            with open(test_input_file, "w") as tif:
                tif.write(test_data)

            # 使用 testinput.txt 作為標準輸入來執行編譯後的程式
            try:
                with open(test_input_file, "r") as tif:
                    run_result = subprocess.run(exe_path,stdin=tif,stdout=subprocess.PIPE,stderr=subprocess.PIPE,text=True,encoding="utf-8",errors="ignore",timeout=test_timeout)    #replace errors="ignore"
            except subprocess.TimeoutExpired:
                msg=f"執行 {os.path.basename(exe_path)} 時逾時 {test_timeout} 秒，已跳過。"
                print(f"{msg}")
                st_info.append(f"{msg}")
                continue
            
            output=run_result.stdout
            if not output.endswith("\n"):
                output += "\n\n"

            # 結果寫入
            retries = 0
            while retries < max_retries:
                try:
                    with open(output_file_path, "a",errors="ignore") as outf:
                        outf.write(output)
                    break
                except PermissionError as e:
                    retries += 1
                    msg=f"PermissionError 正在重試 {retries}/{max_retries} ..."
                    print(f"{msg}")
                    st_info.append(f"{msg}")
                    time.sleep(retry_delay)
            else:
                # 如果重試次數用完，依需求決定要怎麼處理，例如記錄錯誤或跳過
                msg=f"無法寫入檔案 {output_file_path}，請確認檔案是否被其他程式使用。"
                print(f"{msg}")
                st_info.append(f"{msg}")
            

            #學生輸出和標準答案比對
            if idx < len(ans_blocks):
                expected = "\n".join(ans_blocks[idx]) + "\n"
                if output.strip() == expected.strip():
                    msg=f"測資 {idx+1} 正確"
                    print(f"{msg}")
                    st_info.append(f"{msg}")
                    ans_check += 1
                else:
                    msg=f"測資 {idx+1} 錯誤\n預期輸出: {expected}\n實際輸出: {output}"
                    print(f"{msg}")
                    st_info.append(f"{msg}")
            else:
                msg=f"沒有找到測試案例 {idx+1} 的標準答案。"
                print(f"{msg}")
                st_info.append(f"{msg}")

            # clear testinput.txt
            with open(test_input_file, "w") as tif:
                tif.write("")
        
        total_score += score[i-1]*(ans_check/total_count)
        msg=f"第 {i} 題 答對 {ans_check}/{total_count} 測資，得分: {score[i-1]*(ans_check/total_count)}\n"
        print(f"{msg}")
        st_info.append(f"{msg}")

        if ans_check != total_count:
            results.append("X")
            error_count += 1
            wrong_questions.append(i)
        else:
            results.append("O")

    total_score = round(total_score, 2) #取小數後兩位
    return total_score, results, error_count, wrong_questions, st_info


#比對答案
def comparison_student_data(folder, num_problems, high_num_problems,base_dir):
    """
    比對資料夾內的 ioutput.txt 與 ians.txt，
    回傳：
      - results: 每題 "O" 或 "X"
      - error_count: 錯誤題數
      - wrong_questions: 所有答錯的題號清單 (list of int)
    """
    results = []
    error_count = 0
    wrong_questions = []  # 記錄哪幾題錯誤
    high_start= num_problems - high_num_problems + 1
    high_error_count = 0
    test_file_path = os.path.join(base_dir, test_file_dir)

    for i in range(1, num_problems + 1):
        output_file = os.path.join(folder, f"{i}output.txt")
        ans_file_path = os.path.join(test_file_path, f"{i}ans.txt")

        if not os.path.exists(output_file):
            print(f"找不到 {i}output.txt")
            results.append("X")
            if i >= high_start:
                high_error_count += 1
            else:
                error_count += 1
            wrong_questions.append(i)
            continue
        if not os.path.exists(ans_file_path):
            print(f"找不到答案檔 {i}ans.txt，題號 {i}")
            results.append("?")
            if i >= high_start:
                high_error_count += 1
            else:
                error_count += 1
            wrong_questions.append(i)
            continue

        with open(output_file, "r", encoding="utf-8", errors="ignore") as f_out:   #replace errors="replace"
            student_output = f_out.read().strip()
        f_out.close()

        with open(ans_file_path, "r", encoding="utf-8") as f_ans:
            correct_output = f_ans.read().strip()
        f_ans.close()

        if student_output == correct_output:
            results.append("O")
        elif i >= high_start:
            print(f"第 {i} 題與答案不符")
            results.append("X")
            high_error_count += 1
            wrong_questions.append(i)
        else:
            print(f"第 {i} 題與答案不符")
            results.append("X")
            error_count += 1
            wrong_questions.append(i)

    return results, error_count, high_error_count, wrong_questions
#寫入E欄(錯誤題目)
def write_errors_to_excel(student, wrong_questions):
    """
    若 wrong_questions 不為空，則在 Excel (StudentList.xlsx) 裡，
    尋找 A 欄 (dir_name) 與 student 相同的列，將
    '第 1,2,3 題錯誤' 寫入 E 欄 (第 5 欄)。
    """
    st_info=[]
    msg=""

    if not wrong_questions:
        # 全對就不寫任何東西
        return

    if not os.path.exists(excel_file):
        print(f"Excel 檔 {excel_file} 不存在，無法更新。")
        return

    wb = load_workbook(excel_file)
    ws = wb.active  # 假設要更新第一個工作表

    # 尋找 A 欄符合 student 的列
    found_row = None
    for row_idx in range(2, ws.max_row + 1):
        dir_name_cell = ws.cell(row=row_idx, column=1).value
        if dir_name_cell == student:
            found_row = row_idx
            break

    if found_row:
        # 組合字串：例如 "第 1,2,3 題錯誤"
        error_str = "第 " + ",".join(str(q) for q in wrong_questions) + " 題錯誤"
        ws.cell(found_row, 5).value = error_str  # 第 5 欄 (E 欄)
        msg=f"\n已在 Excel 中將 {student} 的錯題寫入 E 欄：{error_str}"
        print(f"{msg}")
        st_info.append(f"{msg}")
    else:
        msg=f"未在 Excel 中找到 {student}，無法寫入錯題資訊。"
        print(f"{msg}")
        st_info.append(f"{msg}")

    wb.save(excel_file)
    return st_info
#寫入C、D欄(錯誤題數、分數)
def add_excel(student,error_count,total_score):
    """
    參數:學生/總題數/高分題數/基本錯題/基本配分/高分錯題/高分配分
    開啟已存在的 StudentList.xlsx，尋找 A 欄與 student 相同的列，
    若找到則將 error_count 寫入 C 欄，(total_problems - error_count) 寫入 D 欄。
    """
    msg=""
    if not os.path.exists(excel_file):
        print(f"Excel 檔 {excel_file} 不存在，無法更新。")
        return

    # 讀取既有的 Excel
    wb = load_workbook(excel_file)
    ws = wb.active  # 假設要更新第一個工作表
    
    # 從第 2 列開始尋找（跳過標題列）
    found_row = None
    for row in range(2, ws.max_row + 1):
        dir_name_cell = ws.cell(row=row, column=1).value  # A 欄
        if dir_name_cell == student:
            found_row = row
            break

    if found_row:
        # C 欄 (column=3) 寫入錯誤題數
        ws.cell(found_row, 3).value = error_count
        # D 欄 (column=4) 寫入總分數
        ws.cell(found_row, 4).value = total_score
        msg=f"已寫入學生: {student} ，錯誤題數 = {error_count}，分數 = {total_score}"
        print(f"{msg}")
    else:
        msg=f"未在 Excel 中找到 {student}，無法更新。"
        print(f"{msg}")

    wb.save(excel_file)
    return msg

#setting excel format
def format_excel(excel_file,avg_score_a):
    """
    開啟指定的 Excel 檔案，將所有儲存格的字體設定為 size=12，
    並自動調整每一欄的寬度，使內容能夠完整顯示。
    """
    if not os.path.exists(excel_file):
        print(f"找不到檔案：{excel_file}")
        return

    # 讀取既有 Excel
    wb = load_workbook(excel_file)
    ws = wb.active  # 假設要處理第一個工作表

    found_row = None
    for row_idx in range(2, ws.max_row + 2):
        dir_name_cell = ws.cell(row=row_idx, column=1).value
        if not dir_name_cell or str(dir_name_cell).strip() == "":
            found_row = row_idx
            break

    avg_score = f"avg score: "
    avg_num = round(sum(avg_score_a) / len(avg_score_a), 2)
    if found_row:
        ws.cell(found_row, 3).value = avg_score
        ws.cell(found_row, 4).value = avg_num
        print(f"已在 Excel 中將 {avg_score}{avg_num} 寫入 C 欄")
    else:
        print(f"未知錯誤，無法寫入 {avg_score}{avg_num}")

    # 1) 設定字體大小為 12
    for row in ws.iter_rows():
        for cell in row:
            # 如果想強制所有儲存格都設定字體大小，即使它是空值也可
            # 若只想設定有值的儲存格，可以檢查 cell.value
            cell.font = Font(size=12)

    # 2) 模擬「自動欄寬」
    # 逐欄找出該欄最長字串長度，然後設定 column_dimensions
    for col_idx, col_cells in enumerate(ws.iter_cols(min_col=1,
                                                     max_col=ws.max_column,
                                                     min_row=1,
                                                     max_row=ws.max_row), start=1):
        max_length = 0
        for cell in col_cells:
            # 若該儲存格有值，計算其長度
            if cell.value is not None:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        # 取出該欄的字母代號 (A, B, C...)
        col_letter = get_column_letter(col_idx)
        # 設定欄寬：可再自行加減，避免過於擁擠
        ws.column_dimensions[col_letter].width = max_length + 5

    # 存檔
    wb.save(excel_file)
    print(f"已完成對 {excel_file} 的字體與欄寬調整。")
    return avg_score+str(avg_num)


def main():
    global excel_file

    print(f"寫入重試次數: {max_retries}\n每次等待時間(s): {retry_delay}\n每個程式最大執行時間(s): {test_timeout}\n")

    # 詢問使用者本次要執行幾個程式
    unzip= int(input("是否解壓縮檔案(預設為1): ") or 1)
    try:
        num_problems = int(input("檢測程式數量? "))
    except ValueError:
        print("請輸入有效的整數。")
        return
    
    score=[]
    avg_score_a=[]
    for i in range(num_problems):
        score.append(int(input(f"第 {i+1} 題配分: ")))
    print(f"\n總分: {sum(score)}")
    print(f"score: {score}\n")

    score_check=input("分數是否正確(預設為y): ")or "y"
    while score_check.lower() == "n":
        if score_check.lower() == "n":
            score_check_count=int(input("請重新輸入第幾題: "))
            if score_check_count>num_problems or score_check_count<=0:
                print(f"\n請輸入正確的題號!\n")
                continue
            score[score_check_count-1]=int(input(f"第 {score_check_count} 題配分: "))
            print(f"\n總分: {sum(score)}")
            print(f"score: {score}\n")
            score_check=input("分數是否正確(預設為y): ")or "y"



    selection=input("作業編號(eg.02261): ")
    
    
    print("\n\n--------------- START INSPECION ---------------")
    
    now = datetime.now()
    start_time = now.strftime("%Y/%m/%d %H:%M:%S")
    start = time.perf_counter()

    #解壓縮
    if unzip:
        unzip_file(os.getcwd())

    #檢查是否有Excel檔案
    excel_file = f"Score_{selection}.xlsx"
    if os.path.exists(excel_file):
        print(f"{excel_file} is True")
        check_excel = 1
    else:
        print(f"{excel_file} is False")
        check_excel = 0

    #rename
    base_dir = os.getcwd()      #當前目錄
    hw_dir_path = os.path.join(base_dir, hw_dir)  #作業資料夾目錄
    #student_root = find_student_root(hw_dir_path)  #學生資料夾目錄
        
    # 生成 Excel 學生清單
    generate_excel(check_excel,score)

    #重新命名學生資料夾並寫入Excel
    #print(f"hw_dir_path: {hw_dir_path} , check_excel: {check_excel}")
    student_folder_name_excel(hw_dir_path,check_excel)

    print("\n\n--------------- CHECK .cpp FILE ---------------")

    #check non cpp files
    move_non_cpp_folders(hw_dir_path)
    print("\n\n--------------- TESTING ---------------")

    ctf=check_test_files(num_problems)
    if ctf:
        return

    num=0
    items = [os.path.join(hw_dir_path, d) for d in os.listdir(hw_dir_path) if os.path.isdir(os.path.join(hw_dir_path, d))]
    total_file_path = os.path.join(base_dir, total_file_name)
    with open(total_file_path, "a", encoding="utf-8") as total_file:
        for item in items:
            student_info_a=[]
            student_name = os.path.basename(item)
            student_info = f"{student_name}.txt"
            student_info_path=os.path.join(item,student_info)
            
            if os.path.isdir(item):
                st_info = []
                if student_name == error_student_folder:
                    #print(f"跳過 {item}，資料夾名稱為 {error_student_folder}")
                    continue
                num+=1
                title_info=f"{num}.處理資料夾：{student_name}"
                print(f"\n\n{title_info}")
                total_score, results, error_count, wrong_questions,st_info=process_student_folder(item, num_problems,score,base_dir)
                
                student_info_a.append(title_info)
                student_info_a.append(st_info)
                st_info=[]

                #print(f"{num}. 學生 {student_name}")
                #print("wrong_questions: ",wrong_questions)
                st_info=write_errors_to_excel(student_name, wrong_questions)
                
                result_str = " ".join(results)
                line = f"{student_name}----- {result_str} -----共錯 {error_count} 題，得分: {total_score:.2f}"
                avg_score_a.append(total_score)
                #學生/總題數/錯誤題數/總分
                msg=add_excel(student_name,error_count,total_score)

                total_file.write(line)
                total_file.write("\n")

                end_info=f"結果: {line.strip()}\n"
                print(f"{end_info}\n\n")
                
                student_info_a.append(st_info)
                student_info_a.append(msg)
                student_info_a.append(end_info)

                with open(student_info_path, "w", encoding="utf-8") as student_file:
                    for data in student_info_a:
                        if isinstance(data, list):
                            # 如果是列表，將裡面的每個元素各自換行寫入
                            # 同時把 None 轉換成空字串，或過濾掉 None 值
                            cleaned_list = [str(item) if item is not None else "" for item in data]
                            student_file.write("\n".join(cleaned_list))
                            student_file.write("\n")
                        else:
                            # 如果 data 為 None，就轉成空字串再寫入
                            student_file.write(str(data) if data is not None else "")
                            student_file.write("\n")
                txt2pdf(student_name,student_info_path,item)


            
    print("\n\n--------------- END INSPECTION ---------------\n")
    
 
    #設定Excel格式
    avg_score_s=format_excel(excel_file,avg_score_a)

    end = time.perf_counter()
    elapsed = end - start
    minutes = int((elapsed % 3600) // 60)
    seconds = elapsed % 60

    execuition_time = f"執行時間: {end - start:.2f} 秒 ({minutes} 分 {seconds:.2f} 秒)"
    average_time = f"共 {num} 位學生，平均處裡時間: {(end - start)/num:.2f} 秒"
    now = datetime.now()
    end_time = now.strftime("%Y/%m/%d %H:%M:%S")
    start_t = f"開始時間: {start_time}"
    end_t = f"結束時間: {end_time}"

    with open(total_file_path, "a", encoding="utf-8") as total_file:
        total_file.write(f"\n{avg_score_s}\n")
        total_file.write(f"{execuition_time}\n")
        total_file.write(f"{average_time}\n")
        total_file.write(f"{start_t}\n")
        total_file.write(f"{end_t}\n")
    total_file.close()

    print(f"\n\n{execuition_time}")
    print(f"{average_time}\n")
    print(f"{start_t}")
    print(f"{end_t}\n")

if __name__ == "__main__":
    main()
    input("執行完畢，請按 Enter 鍵退出...")